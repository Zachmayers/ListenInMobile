import csv
import pandas as pd
from sqlalchemy import create_engine, types
import pandas as pd
from datetime import datetime, timedelta
from datetime import date
import os
import util
import logging
from openpyxl import load_workbook, Workbook
import ffi_data_pull
import match_ffi_ringba_data
#@@@@@ 
#python3 reports_on_matched_data.py --leadspedia_file={} --

LOG = 'reports_on_matched_date.log'

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s:%(levelname)s:%(name)s:%(message)s",
    handlers=[
        logging.FileHandler(LOG),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

ARGS = {
        # 'leadspedia_file': ['path to a file (in CSV format) for leadspedia', True],
        # 'onelink_file': ['path to a file (in CSV format) for onelink', True],
        # 'ringba_file': ['path to a file (in CSV format) for ringba', False],
        # 'leadspedia_phone_column': ['name of the column in LP file containing Phone numbers', True],
        # 'onelink_phone_column': ['name of the column in onelink file containing Phone numbers', True],
        # 'ringba_phone_column': ['name of the column in ringba file containing Phone numbers', False],
        'from_date':['yyyymmdd', False],
        'to_date':['yyymmdd', False],
       }

LP_LEAD_HEADERS = None

RINGBA_HEADERS  = ['tag:User:Agent_User', 'tag:User:LMX_Lead_ID', 'tag:User:ListID', 'tag:User:SupplierID', 
                   'tag:User:source_lead_id', 'tag:CallLength:Total', 'Connected', 'Converted',
                   'Campaign', 'Publisher', 'Is duplicate', 'Target', 'Recording', 'Caller ID']

FFI_HEADERS     = ['Timestamp', 'Generation Method Description', 'Source', 'Lead Type Name', 'Lead_ID',
                   'Day_Phone', 'Notes', 'Result', 'Sold', 'Total Annual Premium', 'State', 'Timestamp Day']

def trim_header_list(headers, file):
    # Returns a reduced list of headers based off of the populated
    # columns in the file given.

    trimmed = []
    df_file = pd.read_csv(file, low_memory=False, dtype=str, encoding="latin-1")

    # This is mostly for LP headers 
    # ~ Used when there is no headerlist given or the list is null.
    # ~ removes columns that are not populated.
    if headers is None:
        for col in df_file.columns:
            if df_file[col].isnull().values.any() != df_file.count()[col]:
                trimmed.append(col)
        return trimmed

    # Otherwise, remove headers from the header list that are not found in the file.
    for header in headers:
        if header in df_file.columns:
            trimmed.append(header)
    
    return trimmed

# Formats RINGBA phone column by removing 1 from beggining.
def remove_1_on_phone(file, column):
    counter = 0
    final_results = []
    with open(file, newline='', encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        header = next(reader, None)
        if header != None:
            for csv_line in reader:
                if csv_line:
                    counter = counter + 1
                    phone = csv_line[header.index(column)]          
                   # logger.info("Processing Record #{}.".format(counter))
                    result = phone.startswith('1')
                    if result and len(phone) > 10:
                        phone = phone[1:]
                        logger.info(f"Changed Phone to {phone}.")
                        csv_line[header.index(column)] = phone
   
                    final_results.append(csv_line)

# Formats FFI Timestamp column to not include time of day
def format_timestamp(file, column):
    counter = 0
    final_results = []
    with open(file, newline='', encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        header = next(reader, None)
        if header != None:
            for csv_line in reader:
                if csv_line:
                    counter = counter + 1
                    timestamp = csv_line[header.index(column)]          
                    # logger.info("Processing Record #{}.".format(counter))
                    # datetime.utcnow().strftime('%d/%m/%Y')
                    i = 0
                    slash = 0
                    while(i < len(str(timestamp))):
                        if timestamp[i] == '/':
                            slash += 1
                        if slash == 2:
                            i += 5
                            break
                        i += 1
                
                    timestamp = timestamp[:i]
                    # logger.info(f"Changed timestamp to {timestamp}.")
                    csv_line[header.index(column)] = timestamp
   
                    final_results.append(csv_line)
    logger.info(f"Formatted {counter} timestamp columns")


def reports_on_matched_data(onelink_file: str):
    dir_path = os.path.dirname(os.path.realpath(__file__))
    date = datetime.strftime(datetime.now(), "%m01") + "-" + (datetime.strftime(datetime.now() - timedelta(days = 1), "%m%d") )
    logger.info(f'Started execution of script  - {date}')
    onelink_phone_column = "Day_Phone"
    ringba_phone_column = "Caller ID"
    #Onelink File
    if onelink_file != None:
        # format_timestamp(onelink_file, 'Timestamp')
        OneLink_df = pd.read_csv(onelink_file, sep=',', usecols = trim_header_list(FFI_HEADERS, onelink_file), low_memory=False, dtype=str, encoding="latin-1")
        # OneLink_df[onelink_phone_column] = OneLink_df[onelink_phone_column].str[:-1]
        # print(datetime.strptime(OneLink_df['Timestamp'],'%m/%d/%Y %H:%M:%S %p').date()
        # OneLink_df['Timestamp'] = [datetime.strptime(d, "%m/%d/%Y %H:%M:%S %p").date() for d in OneLink_df["Timestamp"]]
        # OneLink_df = OneLink_df.add_prefix('OneLink_')
        # onelink_phone_column    = 'OneLink_' + onelink_phone_column
    
    #LeadsPedia file
    # if leadspedia_file != None:
    #     LP_df = pd.read_csv(leadspedia_file, sep=',', usecols= trim_header_list(LP_LEAD_HEADERS, leadspedia_file), low_memory=False, dtype=str, encoding="latin-1")
        # LP_df = LP_df.add_prefix('Leadspedia_')
        # leadspedia_phone_column = 'Leadspedia_' + leadspedia_phone_column

    RINGBA_df = match_ffi_ringba_data.match_data(OneLink_df[onelink_phone_column].to_list())

    # # merge LP files if exists
    # if leadspedia_file:
    #     merged1 = OneLink_df.merge(LP_df, how="left", left_on = onelink_phone_column, right_on =  leadspedia_phone_column)
    #     name1 = 'FFI-LP Data ' + date
    #     # path = os.path.join(dir_path, name)e
    #     # merged1.to_csv(path, index=False)
    #     logger.info('Merged FFI - LP')
    #     logger.info('Matched ' + str(merged1.shape[0]) + ' rows ')
    # # merge Ringba files if exists
    # if ringba_file:
    merged2 = OneLink_df.merge(RINGBA_df, how="left", left_on = onelink_phone_column, right_on = ringba_phone_column)
    name2 = 'FFI-Ringba Data ' + date
    # path = os.path.join(dir_path, name)
    # merged.to_csv(path, index=False)
    logger.info('Merged FFI - Ringba')
    logger.info('Matched ' + str(merged2.shape[0]) + ' rows ')

    excel_name = 'FFI - ' + date + 'Ringba LP.xlsx'
    report_file = os.path.join(dir_path,excel_name)
    logger.info(f'Intiallized excel file {excel_name}')
    if not os.path.isfile(report_file):
        wb = Workbook()
        wb.save(report_file)

    book = load_workbook(report_file)
    writer = pd.ExcelWriter(report_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    merged1.to_excel(writer, sheet_name=name1)
    logger.info(f'Added sheet {name1}')
    merged2.to_excel(writer, sheet_name=name2)
    logger.info(f'Added sheet {name2}')


    writer.save()
    
    

if __name__ == '__main__':
    parser = util.create_arg_parser(ARGS)
    args = parser.parse_args()
    logger.info(str(args))
    path = ffi_data_pull.run_script(args.from_date, args.to_date)
    reports_on_matched_data( path)
    logger.info('Finished execution of script')
    